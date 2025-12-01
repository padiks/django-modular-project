# apps/excel/views.py
from django.shortcuts import render
import os
from django.conf import settings
from openpyxl import load_workbook

def index(request):
    # /home/user/Public/web/project_folder/data/excel.xlsx
    file_path = os.path.join(settings.BASE_DIR, 'data', 'excel.xlsx')  # project_folder/data/excel.xlsx
    xlsx_data = []

    try:
        wb = load_workbook(filename=file_path)
        ws = wb.active  # first sheet

        # Assuming first row is headers, start from row 2
        for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
            # map your columns to keys
            xlsx_data.append({
                "id": row[0],
                "member": row[1],
                "targets": row[2],
                "versus": row[3],
                "total": row[4],
            })

        context = {
            "title": "Excel Data",
            "xlsx": xlsx_data,
        }

    except Exception as e:
        context = {
            "title": "Excel Data",
            "xlsx": [],
            "error": f"Failed to read Excel file: {e}"
        }

    return render(request, 'excel/index.html', context)
